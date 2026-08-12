"""端到端验证脚本（模板标记取值 + 用户表格模板输出，目录助手）。

构造测试模板（矩形区域 + [图号]/[管段编号] 占位符）、三张测试图纸
（有管段 / 无管段→NA / 无图号→文件名兜底）与表格模板
（表头列名 = 字段名 + 页码），跑完整 pipeline 并断言：
  1. 表格模板必填（缺模板报错）
  2. 输出按用户表格模板表头列名填值（动态列、样式保留）
  3. 区域取值正确、无管段填 NA、图号兜底、页码每文件一页
  4. 表头行由占位符字段名反推：表头在第 3 行（前两行为标题）仍正确
  5. 表头与字段名完全不匹配时报错
  6. sheet 自动定位：表头在第二个 sheet（第一个为无关说明页）仍正确
  7. 多个 sheet 表头一致时：自动定位取第一个；指定 sheet_name 用指定 sheet

用法：uv run python scripts/verify_end_to_end.py
"""

from __future__ import annotations

import pathlib
import sys
import tempfile

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import ezdxf  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from cadbatchassistant.core import catalog_excel_writer  # noqa: E402
from cadbatchassistant.core.catalog_pipeline import run_pipeline  # noqa: E402

TMP = pathlib.Path(tempfile.mkdtemp(prefix="cad_verify_"))


def make_template() -> pathlib.Path:
    """图纸模板：矩形A(0,0,20,10) 内 [管段编号]、单点 [图号](30,40)。"""
    doc = ezdxf.new("R2013")
    m = doc.modelspace()
    m.add_lwpolyline([(0, 0), (20, 0), (20, 10), (0, 10)], close=True)
    m.add_text("[管段编号]", dxfattribs={"insert": (5, 4), "height": 1.0})
    m.add_text("[图号]", dxfattribs={"insert": (30, 40), "height": 1.0})
    p = TMP / "模板.dxf"
    doc.saveas(p)
    return p


def make_xlsx_template(header_row: int = 1) -> pathlib.Path:
    """表格模板：表头列名 = 字段名 + 页码（含样式：表头加粗/底色）。

    header_row：表头所在行（1 基）；其上方行填入标题占位内容，用于验证
    表头行由占位符字段名反推（而非固定行号）。
    """
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "目录"
    headers = ["管段编号", "图号", "页码"]
    for r in range(1, header_row):
        ws.cell(row=r, column=1, value=f"标题/公司名 {r}")
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=ci, value=h)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=ci)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFD966")
    p = TMP / f"目录样式模板_表头第{header_row}行.xlsx"
    wb.save(p)
    return p


def make_drawings() -> tuple[pathlib.Path, pathlib.Path, pathlib.Path]:
    """图纸A：有管段+图号；图纸B：无管段→NA；图纸C：无图号→兜底。"""
    d1 = ezdxf.new("R2013")
    m1 = d1.modelspace()
    m1.add_text("PIPE-001-AA", dxfattribs={"insert": (2, 2), "height": 1.0})
    m1.add_text("PIPE-002-BB", dxfattribs={"insert": (8, 6), "height": 1.0})
    m1.add_text("DW-1001", dxfattribs={"insert": (30, 40), "height": 1.0})
    f1 = TMP / "图纸A.dxf"
    d1.saveas(f1)

    d2 = ezdxf.new("R2013")
    m2 = d2.modelspace()
    m2.add_text("DW-1002", dxfattribs={"insert": (30, 40), "height": 1.0})
    f2 = TMP / "图纸B.dxf"
    d2.saveas(f2)

    d3 = ezdxf.new("R2013")
    m3 = d3.modelspace()
    m3.add_text("PIPE-009-XX", dxfattribs={"insert": (2, 2), "height": 1.0})
    f3 = TMP / "DRAW-5566.dxf"
    d3.saveas(f3)
    return f1, f2, f3


def main() -> int:
    failures: list[str] = []
    template = make_template()
    xlsx_tpl = make_xlsx_template()
    f1, f2, f3 = make_drawings()
    out = TMP / "目录.xlsx"

    # 1.表格模板必填
    logs: list[str] = []
    res_no_tpl = run_pipeline(
        template, "", [f1], TMP / "x.xlsx",
        rules={"figure_field": "图号"}, log=logs.append, progress=lambda p: None)
    if res_no_tpl.ok or "表格模板" not in res_no_tpl.error:
        failures.append(f"表格模板必填校验失效: {res_no_tpl.error}")

    # 2. 完整流程
    res = run_pipeline(
        template, xlsx_tpl, [f1, f2, f3], out,
        rules={"figure_field": "图号"},
        log=logs.append, progress=lambda p: None,
    )
    if not res.ok:
        print(f"pipeline 失败: {res.error}")
        return 1
    print(f"[1] 字段列 = {res.fields}，图纸 {res.total_files} 张，NA {res.na_rows} 行，总页 {res.total_pages}")
    if res.fields != ["管段编号", "图号"]:
        failures.append(f"字段列不符: {res.fields}")
    if res.na_rows != 1:
        failures.append(f"NA 行数不符: {res.na_rows}")

    ws = load_workbook(out)["目录"]
    rows = [tuple(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                           values_only=True)]
    for r in rows:
        print("   ", r)
    if rows[0] != ("管段编号", "图号", "页码"):
        failures.append(f"表头不符（应按模板表头）: {rows[0]}")
    if rows[1][0] != "PIPE-001-AA" or rows[2][0] != "PIPE-002-BB":
        failures.append("图纸A 管段取值不符")
    if rows[1][1] != "DW-1001":
        failures.append("图纸A 图号不符")
    na_row = next((r for r in rows if r[1] == "DW-1002"), None)
    if na_row is None or na_row[0] != "NA":
        failures.append("图纸B 管段应填 NA")
    fb_row = next((r for r in rows if r[0] == "PIPE-009-XX"), None)
    if fb_row is None or fb_row[1] != "DRAW-5566":
        failures.append("图纸C 图号应兜底为文件名")
    pages = sorted({r[2] for r in rows[1:] if r[2] is not None})
    if pages != [3, 4, 5]:
        failures.append(f"页码不符: {pages}")

    # 3. 样式保留（模板表头底色 FFD966 应保留到输出）
    head_fill = load_workbook(out)["目录"].cell(row=1, column=1).fill
    if head_fill.fgColor.rgb not in ("00FFD966", "FFFFD966"):
        failures.append(f"模板表头样式未保留: {head_fill.fgColor.rgb}")

    # 3.5 表头行反推：表头不在第 1 行（前两行为标题/公司名）也应正确
    xlsx_tpl3 = make_xlsx_template(header_row=3)
    out3 = TMP / "目录_h3.xlsx"
    res3 = run_pipeline(
        template, xlsx_tpl3, [f1, f2, f3], out3,
        rules={"figure_field": "图号"},
        log=logs.append, progress=lambda p: None,
    )
    if not res3.ok:
        failures.append(f"表头第3行流程失败: {res3.error}")
    else:
        ws3 = load_workbook(out3)["目录"]
        rows3 = [tuple(r) for r in ws3.iter_rows(min_row=1, max_row=ws3.max_row,
                                                 values_only=True)]
        if rows3[2] != ("管段编号", "图号", "页码"):
            failures.append(f"表头第3行反推失败: {rows3[2]}")
        if rows3[3][0] != "PIPE-001-AA":
            failures.append("表头第3行数据错位")
        if rows3[0][0] != "标题/公司名 1" or rows3[1][0] != "标题/公司名 2":
            failures.append("表头上方标题行未保留")

    # 3.6 表头反推失败：模板表头与字段名完全不匹配 → 报错
    wb_bad = Workbook()
    ws_bad = wb_bad.active
    ws_bad.title = "目录"
    ws_bad.append(["名称", "备注", "说明"])
    bad = TMP / "无匹配表头模板.xlsx"
    wb_bad.save(bad)
    res_bad = run_pipeline(
        template, bad, [f1], TMP / "bad.xlsx",
        rules={"figure_field": "图号"},
        log=logs.append, progress=lambda p: None,
    )
    if res_bad.ok or "表头" not in res_bad.error:
        failures.append(f"表头反推失败未报错: {res_bad.error}")

    # 3.7 sheet 自动定位：表头在第二个 sheet（第一个 sheet 为无关说明页）
    wb_multi = Workbook()
    ws_other = wb_multi.active
    ws_other.title = "说明页"
    ws_other.append(["这是说明，无表头"])
    ws_real = wb_multi.create_sheet("目录表")
    for ci, h in enumerate(["管段编号", "图号", "页码"], start=1):
        ws_real.cell(row=1, column=ci, value=h)
    multi = TMP / "多sheet模板.xlsx"
    wb_multi.save(multi)
    out_m = TMP / "目录_multi.xlsx"
    res_m = run_pipeline(
        template, multi, [f1], out_m,
        rules={"figure_field": "图号"},
        log=logs.append, progress=lambda p: None,
    )
    if not res_m.ok:
        failures.append(f"多sheet自动定位失败: {res_m.error}")
    else:
        rows_m = [tuple(r) for r in load_workbook(out_m)["目录表"].iter_rows(
            min_row=1, max_row=load_workbook(out_m)["目录表"].max_row,
            values_only=True)]
        if rows_m[0] != ("管段编号", "图号", "页码") \
                or rows_m[1][0] != "PIPE-001-AA":
            failures.append(f"多sheet自动定位填值不符: {rows_m}")

    # 3.8 并列 sheet：两个表头一致的 sheet——自动定位取第一个；
    #     指定 sheet_name 时用指定 sheet
    wb_tie = Workbook()
    wb_tie.active.title = "目录A"
    wb_tie.active.append(["管段编号", "图号", "页码"])
    ws_tb = wb_tie.create_sheet("目录B")
    ws_tb.append(["管段编号", "图号", "页码"])
    tie = TMP / "并列sheet模板.xlsx"
    wb_tie.save(tie)
    res_t = run_pipeline(
        template, tie, [f1], TMP / "目录_tie.xlsx",
        rules={"figure_field": "图号"},
        log=logs.append, progress=lambda p: None,
    )
    if not res_t.ok:
        failures.append(f"并列sheet流程失败: {res_t.error}")
    else:
        wb_t_out = load_workbook(TMP / "目录_tie.xlsx")
        if wb_t_out.sheetnames != ["目录A", "目录B"]:
            failures.append(f"并列sheet输出 sheet 名不符: {wb_t_out.sheetnames}")
        if wb_t_out["目录A"].cell(row=2, column=1).value != "PIPE-001-AA":
            failures.append("并列sheet自动定位未取第一个 sheet")
    res_tb = run_pipeline(
        template, tie, [f1], TMP / "目录_tieB.xlsx",
        rules={"figure_field": "图号"}, sheet_name="目录B",
        log=logs.append, progress=lambda p: None,
    )
    if not res_tb.ok:
        failures.append(f"指定sheet流程失败: {res_tb.error}")
    else:
        wb_b_out = load_workbook(TMP / "目录_tieB.xlsx")
        if wb_b_out["目录B"].cell(row=2, column=1).value != "PIPE-001-AA":
            failures.append("指定 sheet_name 未写入指定 sheet")

    # 4. write_style_template 可生成参考模板
    ref = TMP / "参考样式模板.xlsx"
    catalog_excel_writer.write_style_template(ref)
    if not ref.is_file():
        failures.append("write_style_template 未生成文件")

    if failures:
        print("\n验证失败：")
        for f in failures:
            print("  -", f)
        return 1
    print("\n全部验证通过 (PASS)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
