"""目录 Excel 输出（文件粒度 + 模板动态列，目录助手）。

内置简单模板：表头 = 模板字段名（出现顺序）+ 「页码」列；
每个 DWG 文件一个条目，行数 = 多值字段最大取值数（至少 1 行）；
无值字段填 NA；单值字段（值数为 1）跨该文件行合并；页码每文件一页合并。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cadbatchassistant.core.catalog_builder import (
    NA,
    Catalog,
    entry_rows,
)

# ---- 内置样式 ----
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")   # 深蓝表头
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN_SIDE = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                 top=_THIN_SIDE, bottom=_THIN_SIDE)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def write_style_template(path: str | Path, fields: list[str] | None = None) -> None:
    """生成可编辑的样式模板 Excel（表头示例 + 数据示例 + 使用说明）。

    fields : 表头列名列表（默认示例 ["字段名", "字段名", "页码"]），
    程序按表头列名与占位符字段名反推表头行，故生成时传入实际字段名
    可保证反推命中。
    用户可编辑此文件自定义表头/数据样式，并在「目录助手」中把它
    选为表格模板；表头行由图纸模板 [字段名] 占位符自动定位。
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "样式模板"
    headers = list(fields) if fields else ["字段名", "字段名", "页码"]
    ws.append(headers)
    ws.append(["示例值" if h != "页码" else 1 for h in headers])
    ws.append([])
    ws.append(["说明", "第一行为表头样式，第二行为数据行样式（含边框/字体/对齐/底色）。"])
    ws.append(["使用", "编辑后保存，在「目录助手」中把它选为表格模板；表头行由图纸模板字段自动定位。"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    for cell in ws[2]:
        cell.border = _BORDER
        cell.alignment = _LEFT
    ws.freeze_panes = "A2"
    for ci, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = 8 if h == "页码" else 24
    wb.save(path)


def detect_header_row(ws, fields: list[str], max_scan: int = 10) -> int | None:
    """反推表格模板的表头行（1 基行号）。

    判定依据 = 图纸模板 [字段名] 占位符解析出的字段名列表：扫描工作表前
    max_scan 行，取与字段名（单元格 strip 后精确匹配）匹配数最多的行作为
    表头行；没有任何行命中字段名时返回 None。
    """
    best_row: int | None = None
    best_score = 0
    for r in range(1, min(int(max_scan), ws.max_row) + 1):
        cells = [str(c.value).strip() if c.value is not None else ""
                 for c in ws[r]]
        score = sum(1 for f in fields if f in cells)
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score > 0 else None


def detect_sheet_candidates(wb, fields: list[str],
                            max_scan: int = 10) -> list[tuple[int, object, int]]:
    """按字段命中数降序返回表格模板中所有候选 sheet。

    返回 [(score, sheet, header_row), ...]：只包含至少命中一个字段的 sheet；
    每个候选的表头行由 detect_header_row 反推。score 相同（并列）时按 sheet
    出现顺序排列，结果稳定可预期——GUI 据此识别并列供用户选择。
    """
    hits: list[tuple[int, object, int]] = []
    for ws in wb.worksheets:
        hr = detect_header_row(ws, fields, max_scan=max_scan)
        if hr is None:
            continue
        cells = [str(c.value).strip() if c.value is not None else ""
                 for c in ws[hr]]
        score = sum(1 for f in fields if f in cells)
        hits.append((score, ws, hr))
    hits.sort(key=lambda t: t[0], reverse=True)
    return hits


def detect_sheet(wb, fields: list[str],
                 max_scan: int = 10) -> tuple[object, int] | None:
    """自动定位表格模板中与字段匹配的 sheet 及其表头行。

    取字段命中数最多的 (sheet, header_row)；无任何匹配返回 None；
    并列最高时返回第一个（靠前）sheet。
    """
    candidates = detect_sheet_candidates(wb, fields, max_scan=max_scan)
    if not candidates:
        return None
    return candidates[0][1], candidates[0][2]


def write_catalog_from_template(catalog: Catalog, xlsx_template: str | Path,
                                out_path: str | Path,
                                sheet_name: str | None = None) -> None:
    """以用户提供的表格模板生成目录，完全保留模板样式。

    sheet_name : 指定使用的 sheet 名（可空）。为空时自动定位：遍历全部
    sheet，取与字段名匹配数最多的 sheet（并列取第一个）。
    表头行定位：用 catalog.fields（图纸模板 [字段名] 占位符解析出的字段名
    列表）反推——扫描目标 sheet 前 10 行，取与字段名匹配数最多的行作为表头行；
    没有任何行匹配字段名时抛 ValueError（提示表头列名应与占位符一致）。
    表头行上方的内容（标题、公司名等）与样式原样保留；数据从表头行下一行
    起，列名 = 模板字段名（与占位符一致）或「页码」；无值的字段填 NA；
    「页码」列填每文件页码。数据区内原有的合并单元格会被取消，由程序按
    单值列/页码列重新合并；图号类列在相邻文件图号相同时跨文件合并为
    一个单元格。
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(xlsx_template))
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"表格模板中不存在名为 {sheet_name} 的 sheet: {wb.sheetnames}")
        ws = wb[sheet_name]
        header_row = detect_header_row(ws, catalog.fields)
        if header_row is None:
            raise ValueError(
                f"表格模板 sheet「{sheet_name}」中未找到与字段匹配的表头行"
                f"（字段：{'、'.join(catalog.fields)}）。"
                "表头列名应包含与图纸模板 [字段名] 占位符一致的字段名。")
    else:
        hit = detect_sheet(wb, catalog.fields)
        if hit is None:
            raise ValueError(
                "表格模板中未找到与字段匹配的表头（sheet 与表头行）（字段："
                + "、".join(catalog.fields)
                + "）。表头列名应包含与图纸模板 [字段名] 占位符一致的字段名。")
        ws, header_row = hit
    first_data_row = header_row + 1
    headers = [str(c.value).strip() if c.value is not None else ""
               for c in ws[header_row]]
    n_cols = len(headers)
    page_col = headers.index("页码") if "页码" in headers else None
    # 图号类列（如 图纸号）与页码列：数据居中显示
    fig_cols = {ci for ci, col in enumerate(headers) if "图" in col and "号" in col}

    # 取消数据区（表头行+1 起、字段列内）原有的合并单元格：
    # MergedCell.value 只读，不先取消会导致下方清空/写入时报
    # AttributeError；单值列与页码列的合并由后续逻辑按程序规则重建。
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        if max_row >= first_data_row and min_col <= n_cols:
            ws.unmerge_cells(str(mr))

    # 清空数据区的值，保留样式
    for r in range(first_data_row, ws.max_row + 1):
        for ci in range(1, n_cols + 1):
            ws.cell(row=r, column=ci).value = None

    row = first_data_row
    # 图号类列：记录每个文件的图号值与行段，供下方「相邻相同图号跨文件合并」
    fig_col_ranges: dict[int, list[tuple[str | None, int, int]]] = {}
    for idx, entry in enumerate(catalog.entries):
        n_rows = entry_rows(entry, catalog.fields)
        page = catalog.total_pages - len(catalog.entries) + idx + 1
        entry_start = row
        for i in range(n_rows):
            for ci, col in enumerate(headers):
                if not col:
                    continue
                if page_col is not None and ci == page_col:
                    cell = ws.cell(row=row, column=ci + 1, value=page)
                    cell.alignment = _CENTER
                    continue
                if col not in catalog.fields:
                    continue  # 非字段列（如备注/版本号）留空
                vals = entry.values.get(col, [])
                if vals:
                    v = vals[i] if i < len(vals) else ""
                    cell = ws.cell(row=row, column=ci + 1, value=v)
                else:
                    cell = ws.cell(row=row, column=ci + 1, value=NA)
                if ci in fig_cols:
                    cell.alignment = _CENTER
            row += 1
        # 单值字段跨行合并（图号类列除外：改由下方按相邻相同值跨文件合并）
        for ci, col in enumerate(headers):
            if not col or ci in fig_cols:
                continue
            if len(entry.values.get(col, [])) == 1 and n_rows > 1:
                ws.merge_cells(
                    start_row=entry_start, start_column=ci + 1,
                    end_row=row - 1, end_column=ci + 1)
        # 页码列跨文件行合并
        if page_col is not None and n_rows > 1:
            ws.merge_cells(
                start_row=entry_start, start_column=page_col + 1,
                end_row=row - 1, end_column=page_col + 1)
        # 记录图号类列的取值与行段（仅单值参与跨文件合并）
        for ci, col in enumerate(headers):
            if not col or ci not in fig_cols:
                continue
            vals = entry.values.get(col, [])
            fig_col_ranges.setdefault(ci, []).append(
                (vals[0] if len(vals) == 1 else None, entry_start, row))

    # 图号类列：相邻文件图号相同 → 跨文件合并为一个单元格
    for ci, ranges in fig_col_ranges.items():
        i = 0
        while i < len(ranges):
            val, s0, _ = ranges[i]
            if val is None:
                i += 1
                continue
            j = i + 1
            while j < len(ranges) and ranges[j][0] == val:
                j += 1
            end = ranges[j - 1][2]
            if end - s0 > 1:
                ws.merge_cells(
                    start_row=s0, start_column=ci + 1,
                    end_row=end - 1, end_column=ci + 1)
            i = j

    wb.save(out_path)
