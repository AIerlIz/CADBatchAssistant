"""读取数据表（.xlsx / .xls），输出 {图纸文件名(不含扩展名): {列名: 值}}。

- 第一行作为表头，后续每行对应一张图纸，第一列为 CAD 文件名。
- DATE 列：Excel 日期序列号（如 46244）转成 YYYY/MM/DD 字符串。
- 重复列名（如两个 REV）自动加 _2 后缀区分。
- 按扩展名分发：.xls → xlrd（旧版 BIFF），其余（.xlsx/.xlsm）→ openpyxl。
"""

from __future__ import annotations

import datetime
from pathlib import Path

from cadbatchassistant.core.common.filetypes import CAD_SUFFIXES

EXCEL_EPOCH = datetime.date(1899, 12, 30)


def _serial_to_date(value) -> str | None:
    """把 Excel 日期序列号或 datetime 转成 YYYY/MM/DD；无法转换返回 None。"""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, (int, float)):
        try:
            d = EXCEL_EPOCH + datetime.timedelta(days=float(value))
            return d.strftime("%Y/%m/%d")
        except (OverflowError, ValueError):
            return None
    return None


def _make_cols(header: list) -> list[str]:
    """表头 → 列名列表（重复列名加 _2 后缀，与解析逻辑一致）。"""
    seen: dict[str, int] = {}
    cols: list[str] = []
    for h in header:
        name = str(h).strip() if h is not None else ""
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        cols.append(name)
    return cols


def load_sheet_meta(path: str | Path) -> tuple[list[str], dict[str, list[str]]]:
    """一次打开数据表，返回 (工作表名列表, {工作表名: 首行表头列表})。

    供 GUI 下拉刷新使用：一次解析即得工作表名与各表首行表头，
    避免多次全量加载（大表只解析一次）。空工作表（无首行）表头记为 []。
    """
    p = str(path)
    if p.lower().endswith(".xls") and not p.lower().endswith(".xlsx"):
        import xlrd

        book = xlrd.open_workbook(p)
        names = list(book.sheet_names())
        headers: dict[str, list[str]] = {}
        for name in names:
            ws = book.sheet_by_name(name)
            headers[name] = _make_cols(ws.row_values(0)) if ws.nrows > 0 else []
        return names, headers
    import openpyxl

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        headers = {}
        for name in names:
            ws = wb[name]
            row = next(ws.iter_rows(values_only=True, max_row=1), None)
            headers[name] = _make_cols(list(row)) if row is not None else []
        return names, headers
    finally:
        wb.close()


def _read_raw_rows(
    path: str | Path, sheet: str | None = None
) -> tuple[list[list], str]:
    """读取指定工作表的原始行列表；返回 (rows, 后端类型 'xlsx'|'xls')。

    sheet 为工作表名，None 时取第一个（向后兼容）。
    """
    p = str(path)
    if p.lower().endswith(".xls") and not p.lower().endswith(".xlsx"):
        return _read_raw_rows_xls(p, sheet), "xls"
    return _read_raw_rows_xlsx(p, sheet), "xlsx"


def _read_raw_rows_xlsx(path: str, sheet: str | None = None) -> list[list]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet is None:
            ws = wb.worksheets[0]
        else:
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"数据表中不存在工作表「{sheet}」，可用："
                    + ("、".join(wb.sheetnames) if wb.sheetnames else "(无)")
                )
            ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return rows


def _read_raw_rows_xls(path: str, sheet: str | None = None) -> list[list]:
    import xlrd

    book = xlrd.open_workbook(path)
    if sheet is None:
        ws = book.sheet_by_index(0)
    else:
        try:
            ws = book.sheet_by_name(sheet)
        except xlrd.biffh.XLRDError:
            names = book.sheet_names()
            raise ValueError(
                f"数据表中不存在工作表「{sheet}」，可用："
                + ("、".join(names) if names else "(无)")
            ) from None
    rows: list[list] = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except (ValueError, OverflowError):
                    row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def get_headers(path: str | Path, sheet: str | None = None) -> list[str]:
    """返回指定工作表首行列名列表（与解析列名逻辑一致）。"""
    rows, _ = _read_raw_rows(path, sheet)
    if not rows:
        raise ValueError(f"空表: {path}")
    return _make_cols(rows[0])


def _build_records(
    raw_rows: list[list], path: str, match_col: str | None = None
) -> dict[str, dict[str, str]]:
    """由原始行列表构建 {图纸名: {列名: 值}}；首行为表头。

    match_col：指定哪一列作为图纸名（匹配键）；None 时默认第一列（向后兼容）。
    """
    if not raw_rows:
        raise ValueError(f"空表: {path}")
    header = raw_rows[0]
    cols = _make_cols(header)
    if match_col is None:
        key_idx = 0
    else:
        try:
            key_idx = cols.index(match_col)
        except ValueError:
            raise ValueError(
                f"数据表中不存在匹配列「{match_col}」，可用列："
                + ("、".join(cols) if cols else "(空表头)")
            ) from None

    result: dict[str, dict[str, str]] = {}
    for row in raw_rows[1:]:
        if not any(c is not None for c in row):
            continue
        record: dict[str, str] = {}
        for col, raw in zip(cols, row, strict=False):
            if col == "":
                continue
            if raw is None:
                record[col] = ""
                continue
            if col == "DATE" or col.startswith("DATE_"):
                s = _serial_to_date(raw)
                if s is not None:
                    record[col] = s
                    continue
            if isinstance(raw, float) and raw.is_integer():
                record[col] = str(
                    int(raw)
                )  # xlrd 数字为 float，1.0 → '1'（与 openpyxl 一致）
            else:
                record[col] = str(raw).strip()

        name = record.get(cols[key_idx], "")
        if not name:
            continue
        stem = Path(name).stem if name.lower().endswith(CAD_SUFFIXES) else name
        result[stem] = record
    return result


def _load_xlsx(
    path: str, match_col: str | None = None, sheet: str | None = None
) -> dict[str, dict[str, str]]:
    """openpyxl 后端：.xlsx / .xlsm 等。"""
    return _build_records(_read_raw_rows_xlsx(path, sheet), path, match_col)


def _load_xls(
    path: str, match_col: str | None = None, sheet: str | None = None
) -> dict[str, dict[str, str]]:
    """xlrd 后端：旧版 .xls（BIFF）。"""
    return _build_records(_read_raw_rows_xls(path, sheet), path, match_col)


def load_xlsx(
    path: str | Path, match_col: str | None = None, sheet: str | None = None
) -> dict[str, dict[str, str]]:
    """读取数据表（.xlsx/.xlsm 或 .xls），返回 {图纸名: {列名: 值}}。

    match_col：图纸名列（默认第一列，向后兼容）。
    sheet    ：工作表名（默认第一个，向后兼容）。
    """
    p = str(path)
    if p.lower().endswith(".xls") and not p.lower().endswith(".xlsx"):
        return _load_xls(p, match_col, sheet)
    return _load_xlsx(p, match_col, sheet)


def load_xlsx_with_headers(
    path: str | Path, match_col: str | None = None, sheet: str | None = None
) -> tuple[dict[str, dict[str, str]], list[str]]:
    """一次读取数据表，同时返回 (数据, 表头列名列表)。

    流水线同时需要「占位符与表头匹配」（取列名）与「按图取行」（取数据）；
    分开调用 get_headers + load_xlsx 会整表解析两次，大表成本翻倍。
    本函数底部一次 _read_raw_rows 同时产出两者，列名与 _make_cols 完全一致
    （record 的键即列名），保证两处语义不漂移。
    """
    raw_rows, _ = _read_raw_rows(path, sheet)
    if not raw_rows:
        raise ValueError(f"空表: {path}")
    headers = _make_cols(raw_rows[0])
    data = _build_records(raw_rows, path, match_col)
    return data, headers


if __name__ == "__main__":
    import json
    import sys

    p = sys.argv[1] if len(sys.argv) > 1 else r"D:\ISO图\数据表.xlsx"
    data = load_xlsx(p)
    print(json.dumps(data, ensure_ascii=False, indent=2))
