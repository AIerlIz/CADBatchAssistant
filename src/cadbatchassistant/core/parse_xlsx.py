# -*- coding: utf-8 -*-
"""读取数据表（.xlsx / .xls），输出 {图纸文件名(不含扩展名): {列名: 值}}。

- 第一行作为表头，后续每行对应一张图纸，第一列为 CAD 文件名。
- DATE 列：Excel 日期序列号（如 46244）转成 YYYY/MM/DD 字符串。
- 重复列名（如两个 REV）自动加 _2 后缀区分。
- 按扩展名分发：.xls → xlrd（旧版 BIFF），其余（.xlsx/.xlsm）→ openpyxl。
"""

from __future__ import annotations

import datetime
from pathlib import Path

EXCEL_EPOCH = datetime.date(1899, 12, 30)


def _serial_to_date(value) -> str | None:
    """把 Excel 日期序列号或 datetime 转成 YYYY/MM/DD；无法转换返回 None。"""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, (int, float)):
        try:
            d = EXCEL_EPOCH + datetime.timedelta(days=float(value))
            return d.strftime("%Y/%m/%d")
        except (OverflowError, ValueError):
            return None
    return None


def _make_cols(header: list) -> list[str]:
    """表头 → 列名列表（重复列名加 _2 后缀，与解析逻辑一致）。"""
    seen: dict[str, int] = {}
    cols: list[str] = []
    for h in header:
        name = str(h).strip() if h is not None else ""
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        cols.append(name)
    return cols


def _read_raw_rows(path: str) -> tuple[list[list], str]:
    """读取首表原始行列表；返回 (rows, 后端类型 'xlsx'|'xls')。"""
    p = str(path)
    if p.lower().endswith(".xls") and not p.lower().endswith(".xlsx"):
        return _read_raw_rows_xls(p), "xls"
    return _read_raw_rows_xlsx(p), "xlsx"


def _read_raw_rows_xlsx(path: str) -> list[list]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_raw_rows_xls(path: str) -> list[list]:
    import xlrd

    book = xlrd.open_workbook(path)
    ws = book.sheet_by_index(0)
    rows: list[list] = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except (ValueError, OverflowError):
                    row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def get_headers(path: str | Path) -> list[str]:
    """返回数据表首行列名列表（与解析列名逻辑一致）。"""
    rows, _ = _read_raw_rows(path)
    if not rows:
        raise ValueError(f"空表: {path}")
    return _make_cols(rows[0])


def _build_records(raw_rows: list[list], path: str) -> dict[str, dict[str, str]]:
    """由原始行列表构建 {图纸名: {列名: 值}}；首行为表头。"""
    if not raw_rows:
        raise ValueError(f"空表: {path}")
    header = raw_rows[0]
    cols = _make_cols(header)

    result: dict[str, dict[str, str]] = {}
    for row in raw_rows[1:]:
        if not any(c is not None for c in row):
            continue
        record: dict[str, str] = {}
        for col, raw in zip(cols, row):
            if col == "":
                continue
            if raw is None:
                record[col] = ""
                continue
            if col == "DATE" or col.startswith("DATE_"):
                s = _serial_to_date(raw)
                if s is not None:
                    record[col] = s
                    continue
            if isinstance(raw, float) and raw.is_integer():
                record[col] = str(int(raw))  # xlrd 数字为 float，1.0 → '1'（与 openpyxl 一致）
            else:
                record[col] = str(raw).strip()

        name = record.get(cols[0], "")
        if not name:
            continue
        stem = Path(name).stem if name.lower().endswith((".dwg", ".dxf")) else name
        result[stem] = record
    return result


def _load_xlsx(path: str) -> dict[str, dict[str, str]]:
    """openpyxl 后端：.xlsx / .xlsm 等。"""
    return _build_records(_read_raw_rows_xlsx(path), path)


def _load_xls(path: str) -> dict[str, dict[str, str]]:
    """xlrd 后端：旧版 .xls（BIFF）。"""
    return _build_records(_read_raw_rows_xls(path), path)


def load_xlsx(path: str | Path) -> dict[str, dict[str, str]]:
    """读取数据表（.xlsx/.xlsm 或 .xls），返回 {图纸名: {列名: 值}}。"""
    p = str(path)
    if p.lower().endswith(".xls") and not p.lower().endswith(".xlsx"):
        return _load_xls(p)
    return _load_xlsx(p)


if __name__ == "__main__":
    import json
    import sys

    p = sys.argv[1] if len(sys.argv) > 1 else r"D:\ISO图\数据表.xlsx"
    data = load_xlsx(p)
    print(json.dumps(data, ensure_ascii=False, indent=2))
