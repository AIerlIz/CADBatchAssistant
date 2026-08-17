"""catalog_excel_writer 表头反推、样式模板与合并逻辑单测（目录助手）。"""

from openpyxl import Workbook, load_workbook

from cadbatchassistant.core.catalog.catalog_builder import (
    Catalog,
    FileEntry,
    is_fig_no_col,
)
from cadbatchassistant.core.catalog.catalog_excel_writer import (
    detect_header_row,
    detect_sheet,
    detect_sheet_candidates,
    write_catalog_from_template,
    write_style_template,
)


def _ws_with_rows(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    return ws


def test_detect_header_row_first_line():
    ws = _ws_with_rows([["管段编号", "图号", "页码"]])
    assert detect_header_row(ws, ["管段编号", "图号"]) == 1


def test_detect_header_row_after_title_lines():
    ws = _ws_with_rows([["某设计院"], ["项目名称"], ["管段编号", "图号", "页码"]])
    assert detect_header_row(ws, ["管段编号", "图号"]) == 3


def test_detect_header_row_no_match_returns_none():
    ws = _ws_with_rows([["名称", "备注", "说明"]])
    assert detect_header_row(ws, ["管段编号", "图号"]) is None


def test_detect_header_row_picks_best_match_row():
    ws = _ws_with_rows(
        [
            ["图号"],  # 命中 1 个
            ["管段编号", "图号", "页码"],  # 命中 2 个 → 表头
            ["管段编号"],  # 命中 1 个
        ]
    )
    assert detect_header_row(ws, ["管段编号", "图号"]) == 2


def test_detect_header_row_empty_fields():
    ws = _ws_with_rows([["管段编号", "图号"]])
    assert detect_header_row(ws, []) is None


def test_write_style_template_with_fields(tmp_path):
    p = tmp_path / "style.xlsx"
    write_style_template(p, fields=["管段编号", "图号", "页码"])
    ws = load_workbook(p).active
    assert detect_header_row(ws, ["管段编号", "图号"]) == 1


def test_write_style_template_default_headers(tmp_path):
    p = tmp_path / "style.xlsx"
    write_style_template(p)
    ws = load_workbook(p).active
    assert ws.cell(row=1, column=1).value == "字段名"
    assert ws.cell(row=1, column=3).value == "页码"


# ---------------- write_catalog_from_template 合并逻辑 ----------------


def _fig_template(tmp_path):
    """表格模板：表头 = 图纸号/包含管段/页码（第 1 行）。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["图纸号", "包含管段", "页码"])
    p = tmp_path / "tpl.xlsx"
    wb.save(p)
    return p


def test_adjacent_same_fig_merged_across_files(tmp_path):
    """相邻文件图号相同：图号列跨文件合并；页码列仍按文件合并。"""
    cat = Catalog(
        fields=["图纸号", "包含管段"],
        entries=[
            FileEntry(
                filename="A", values={"图纸号": ["DWG-1"], "包含管段": ["P-1", "P-2"]}
            ),
            FileEntry(filename="B", values={"图纸号": ["DWG-1"], "包含管段": ["P-3"]}),
        ],
        page_count=1,
        total_pages=4,
        na_rows=0,
    )
    out = tmp_path / "out.xlsx"
    write_catalog_from_template(cat, _fig_template(tmp_path), out)
    ws = load_workbook(out).active
    # A 2 行 + B 1 行 = 3 行数据；图号列跨文件合并 A2:A4
    assert "A2:A4" in {str(m) for m in ws.merged_cells.ranges}
    assert ws.cell(row=2, column=1).value == "DWG-1"
    assert ws.cell(row=2, column=2).value == "P-1"
    assert ws.cell(row=3, column=2).value == "P-2"
    assert ws.cell(row=4, column=2).value == "P-3"
    # 页码列：A 的 2 行合并为 3，B 独立 4
    assert "C2:C3" in {str(m) for m in ws.merged_cells.ranges}
    assert ws.cell(row=2, column=3).value == 3
    assert ws.cell(row=4, column=3).value == 4


def test_adjacent_diff_fig_not_merged_across_files(tmp_path):
    """相邻文件图号不同：图号列不跨文件合并，文件内仍合并。"""
    cat = Catalog(
        fields=["图纸号", "包含管段"],
        entries=[
            FileEntry(
                filename="A", values={"图纸号": ["DWG-1"], "包含管段": ["P-1", "P-2"]}
            ),
            FileEntry(filename="B", values={"图纸号": ["DWG-2"], "包含管段": ["P-3"]}),
        ],
        page_count=1,
        total_pages=4,
        na_rows=0,
    )
    out = tmp_path / "out2.xlsx"
    write_catalog_from_template(cat, _fig_template(tmp_path), out)
    ws = load_workbook(out).active
    merged = {str(m) for m in ws.merged_cells.ranges}
    assert "A2:A3" in merged  # 文件 A 内部 2 行合并
    assert "A2:A4" not in merged  # 不跨文件
    assert ws.cell(row=2, column=1).value == "DWG-1"
    assert ws.cell(row=4, column=1).value == "DWG-2"


def test_same_fig_not_adjacent_not_merged(tmp_path):
    """图号相同但不相邻（中间隔了其他图号）：不合并。"""
    cat = Catalog(
        fields=["图纸号", "包含管段"],
        entries=[
            FileEntry(filename="A", values={"图纸号": ["DWG-1"], "包含管段": ["P-1"]}),
            FileEntry(filename="B", values={"图纸号": ["DWG-2"], "包含管段": ["P-2"]}),
            FileEntry(filename="C", values={"图纸号": ["DWG-1"], "包含管段": ["P-3"]}),
        ],
        page_count=1,
        total_pages=5,
        na_rows=0,
    )
    out = tmp_path / "out3.xlsx"
    write_catalog_from_template(cat, _fig_template(tmp_path), out)
    ws = load_workbook(out).active
    merged = {str(m) for m in ws.merged_cells.ranges}
    assert "A2:A4" not in merged
    assert ws.cell(row=2, column=1).value == "DWG-1"
    assert ws.cell(row=3, column=1).value == "DWG-2"
    assert ws.cell(row=4, column=1).value == "DWG-1"


# ---------------- detect_sheet / detect_sheet_candidates ----------------


def _multi_sheet_wb():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "说明页"
    ws1.append(["无关内容"])
    ws2 = wb.create_sheet("目录表")
    ws2.append(["管段编号", "图号", "页码"])
    return wb


def test_detect_sheet_second_sheet():
    wb = _multi_sheet_wb()
    hit = detect_sheet(wb, ["管段编号", "图号"])
    assert hit is not None
    ws, hr = hit
    assert ws.title == "目录表" and hr == 1


def test_detect_sheet_candidates_sorted_by_score():
    wb = _multi_sheet_wb()
    cands = detect_sheet_candidates(wb, ["管段编号", "图号"])
    assert [(c[0], c[1].title) for c in cands] == [(2, "目录表")]


def test_detect_sheet_tie_keeps_first_and_order_stable():
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "目录A"
    ws_a.append(["管段编号", "图号", "页码"])
    ws_b = wb.create_sheet("目录B")
    ws_b.append(["管段编号", "图号", "页码"])
    cands = detect_sheet_candidates(wb, ["管段编号", "图号"])
    assert len(cands) == 2
    assert cands[0][0] == cands[1][0] == 2
    assert [c[1].title for c in cands] == ["目录A", "目录B"]
    hit = detect_sheet(wb, ["管段编号", "图号"])
    assert hit[0].title == "目录A"


def test_detect_sheet_no_match():
    wb = Workbook()
    wb.active.title = "无"
    wb.active.append(["名称", "备注"])
    assert detect_sheet(wb, ["管段编号", "图号"]) is None
    assert detect_sheet_candidates(wb, ["管段编号", "图号"]) == []


def test_is_fig_no_col_judgement():
    """图号类列判定：命中 图号/图纸号/图幅号，排除「图例符号」等误判。"""
    assert is_fig_no_col("图号")
    assert is_fig_no_col("图纸号")
    assert is_fig_no_col("图幅号")
    assert not is_fig_no_col("图例符号")
    assert not is_fig_no_col("图形编号")
    assert not is_fig_no_col("包含管段")
    assert not is_fig_no_col("页码")
