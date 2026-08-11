# -*- coding: utf-8 -*-
"""数据表解析（parse_xlsx）测试：匹配列（match_col）选择逻辑。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from cadbatchassistant.core import parse_xlsx


def _make_xlsx(path, header: list, rows: list) -> None:
    """用 openpyxl 写一个临时 .xlsx。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _make_xlsx_multi(path, sheets: dict) -> None:
    """用 openpyxl 写一个含多个工作表的临时 .xlsx。

    sheets: {工作表名: (表头列表, 行列表)}
    """
    import openpyxl

    wb = openpyxl.Workbook()
    first = True
    for name, (header, rows) in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = name
        ws.append(header)
        for r in rows:
            ws.append(r)
    wb.save(path)
    wb.close()


class BuildRecordsMatchColTest(unittest.TestCase):
    """_build_records：匹配列选择的核心逻辑（纯内存）。"""

    def test_default_first_col(self):
        raw = [["图号", "名称"], ["A-1", "图纸1"], ["A-2", "图纸2"]]
        data = parse_xlsx._build_records(raw, "x.xlsx")
        self.assertEqual(set(data), {"A-1", "A-2"})
        self.assertEqual(data["A-1"]["名称"], "图纸1")

    def test_match_col_named(self):
        raw = [["序号", "图号", "名称"], [1, "A-1", "图纸1"]]
        data = parse_xlsx._build_records(raw, "x.xlsx", match_col="图号")
        self.assertEqual(set(data), {"A-1"})
        # 非匹配列仍保留在记录中
        self.assertEqual(data["A-1"]["序号"], "1")
        self.assertEqual(data["A-1"]["名称"], "图纸1")

    def test_match_col_missing_raises(self):
        raw = [["图号", "名称"], ["A-1", "图纸1"]]
        with self.assertRaises(ValueError) as cm:
            parse_xlsx._build_records(raw, "x.xlsx", match_col="不存在的列")
        self.assertIn("不存在匹配列", str(cm.exception))
        self.assertIn("图号", str(cm.exception))  # 提示可用列

    def test_dup_cols_suffix_usable_as_match(self):
        raw = [["REV", "REV"], ["1", "2"]]
        data = parse_xlsx._build_records(raw, "x.xlsx", match_col="REV_2")
        self.assertEqual(set(data), {"2"})
        self.assertEqual(data["2"]["REV"], "1")


class LoadXlsxMatchColTest(unittest.TestCase):
    """load_xlsx：文件级透传 match_col。"""

    def setUp(self):
        self._tmp = Path(tempfile.mkdtemp(prefix="px_test_"))

    def tearDown(self):
        for f in self._tmp.iterdir():
            f.unlink(missing_ok=True)
        self._tmp.rmdir()

    def test_load_default_and_named(self):
        p = self._tmp / "data.xlsx"
        _make_xlsx(p, ["序号", "图号", "NPD (inch)"],
                   [[1, "A-1", 12.5], [2, "A-2", 8.0]])
        by_default = parse_xlsx.load_xlsx(p)
        self.assertEqual(set(by_default), {"1", "2"})  # 第一列「序号」为键
        by_fig = parse_xlsx.load_xlsx(p, match_col="图号")
        self.assertEqual(set(by_fig), {"A-1", "A-2"})
        self.assertEqual(by_fig["A-1"]["NPD (inch)"], "12.5")
        self.assertEqual(by_fig["A-2"]["NPD (inch)"], "8")

    def test_load_missing_col_raises(self):
        p = self._tmp / "data.xlsx"
        _make_xlsx(p, ["图号"], [["A-1"]])
        with self.assertRaises(ValueError):
            parse_xlsx.load_xlsx(p, match_col="BOGUS")


class ListSheetsTest(unittest.TestCase):
    """list_sheets / 按 sheet 读取 / sheet 不存在报错。"""

    def setUp(self):
        self._tmp = Path(tempfile.mkdtemp(prefix="px_sheet_"))

    def tearDown(self):
        for f in self._tmp.iterdir():
            f.unlink(missing_ok=True)
        self._tmp.rmdir()

    def test_list_sheets(self):
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {
            "数据表": (["图号", "名称"], [["A-1", "图1"]]),
            "说明": (["说明"], [["xxx"]]),
        })
        self.assertEqual(parse_xlsx.list_sheets(p), ["数据表", "说明"])

    def test_load_named_sheet(self):
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {
            "数据表": (["图号", "名称"], [["A-1", "图1"], ["A-2", "图2"]]),
            "说明": (["说明"], [["yyy"]]),
        })
        by_named = parse_xlsx.load_xlsx(p, sheet="数据表")
        self.assertEqual(set(by_named), {"A-1", "A-2"})
        # sheet 为 None 时仍取第一个工作表（向后兼容）
        by_default = parse_xlsx.load_xlsx(p)
        self.assertEqual(set(by_default), {"A-1", "A-2"})
        # 与匹配列可同时指定
        by_both = parse_xlsx.load_xlsx(p, sheet="数据表", match_col="名称")
        self.assertEqual(set(by_both), {"图1", "图2"})

    def test_sheet_missing_raises(self):
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {"数据表": (["图号"], [["A-1"]])})
        with self.assertRaises(ValueError) as cm:
            parse_xlsx.load_xlsx(p, sheet="不存在的表")
        self.assertIn("不存在工作表", str(cm.exception))
        self.assertIn("数据表", str(cm.exception))  # 提示可用工作表

    def test_get_headers_with_sheet(self):
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {
            "数据表": (["图号", "名称"], [["A-1", "图1"]]),
            "其他": (["X"], [["1"]]),
        })
        self.assertEqual(parse_xlsx.get_headers(p, sheet="其他"), ["X"])
        self.assertEqual(parse_xlsx.get_headers(p), ["图号", "名称"])


class LoadSheetMetaTest(unittest.TestCase):
    """load_sheet_meta：一次打开同时返回工作表名与各表首行表头（GUI 下拉刷新用）。"""

    def setUp(self):
        self._tmp = Path(tempfile.mkdtemp(prefix="px_meta_"))

    def tearDown(self):
        for f in self._tmp.iterdir():
            f.unlink(missing_ok=True)
        self._tmp.rmdir()

    def test_names_and_headers(self):
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {
            "数据表": (["图号", "名称"], [["A-1", "图1"]]),
            "说明": (["说明"], [["xxx"]]),
        })
        names, headers = parse_xlsx.load_sheet_meta(p)
        self.assertEqual(names, ["数据表", "说明"])
        self.assertEqual(headers["数据表"], ["图号", "名称"])
        self.assertEqual(headers["说明"], ["说明"])

    def test_empty_sheet_header_empty(self):
        """空工作表（无首行）表头记为 []，不影响其他表。"""
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {
            "数据表": (["图号"], [["A-1"]]),
            "空表": ([], []),
        })
        names, headers = parse_xlsx.load_sheet_meta(p)
        self.assertEqual(names, ["数据表", "空表"])
        self.assertEqual(headers["数据表"], ["图号"])
        self.assertEqual(headers["空表"], [])

    def test_consistent_with_list_sheets_and_get_headers(self):
        """与旧接口（list_sheets / get_headers）结果一致，行为等价。"""
        p = self._tmp / "multi.xlsx"
        _make_xlsx_multi(p, {
            "数据表": (["图号", "名称"], [["A-1", "图1"]]),
            "其他": (["X"], [["1"]]),
        })
        names, headers = parse_xlsx.load_sheet_meta(p)
        self.assertEqual(names, parse_xlsx.list_sheets(p))
        for name in names:
            self.assertEqual(headers[name], parse_xlsx.get_headers(p, sheet=name))

    def test_dup_cols_suffix(self):
        """重复列名加 _2 后缀，与 get_headers 一致。"""
        p = self._tmp / "dup.xlsx"
        _make_xlsx(p, ["REV", "REV"], [["1", "2"]])
        _, headers = parse_xlsx.load_sheet_meta(p)
        self.assertEqual(headers["Sheet"], ["REV", "REV_2"])


if __name__ == "__main__":
    unittest.main()
